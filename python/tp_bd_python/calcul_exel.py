import psycopg2
conn=psycopg2.connect('dbname=airdata')
cur=conn.cursor()
from openpyxl import load_workbook
objet = load_workbook(filename = 'aeroports.xlsx')
print(objet.active)
avions=objet['avions']
Aeroport=objet['aeroports']
compagines=objet['compagnies']
vols=objet['vols']
print(avions)
print("il est de type Worksheet")
print(compagines)
#print(vols)
liste=(list(avions.rows))
#for elt in liste:
	#print(elt)
#print((avions.rows))
#il retourne un ensemble d'objet de type cell
feuilles=objet.worksheets
#il retourne un ensemble d'objet de type worksheet 
#print(feuilles)
compagines.title="sarata"
compagnie=list(compagines.rows)
#print(compagines.title)
#print(compagnie)
#print(objet.sheetnames)
#for sheet in objet:
    #print(sheet.title)
def insert_avion(avions):
	cellules_avions=avions.values
	tab=[]
	for i in cellules_avions:

		#print(i[0])
		#print(i)
		#cur.execute("insert into avions(modele,nbre_places) values(%s,%s)",(i[1],i[2]))
		tab.append(i)
	#print(tab)
	for i in range(len(tab)):
		if i!=0:
			tuples=(tab[i])
			print (tuples)
			
		
			cur.execute("insert into avions(modele,nbre_places) values(%s,%s)",(tuples[1],tuples[2]))
			conn.commit()
#insert_avion(avions)
def insert_compagnies(feuille_compagines):
	cellules_compagines=feuille_compagines.values
	tab=[]
	for i in cellules_compagines:
		tab.append(i)
	
	for i in range(len(tab)):
		if i!=0:
			tuples=(tab[i])
			print (tuples)
			
		
			cur.execute("insert into Compagnie(nom,nbre_appareil) values(%s,%s)",(tuples[1],tuples[2]))
			conn.commit()
			
#insert_compagnies(compagines)		
			
			
		

def insert_Aeroport(feuille_Aeroport):
	cellules_Aeroport=feuille_Aeroport.values
	tab=[]
	for i in cellules_Aeroport:
		tab.append(i)
	
	for i in range(len(tab)):
		if i!=0:
			tuples=(tab[i])
			print (tuples)
			
		
			cur.execute("insert into Aeroprot values(%s,%s,%s)",(tuples[0],tuples[1],tuples[2]))
			conn.commit()
		

#insert_Aeroport(Aeroport)

def insert_vols(feuille_vols):
	cellules_vols=feuille_vols.values
	tab=[]
	for i in cellules_vols:
		tab.append(i)
	
	for i in range(len(tab)):
		if i!=0:
			tuples=(tab[i])
			#print (tuples)
			
		
			cur.execute("insert into vols(id_vols,depart,arrivee,nbre_passagers,heure_vol) values(%s,%s,%s,%s,%s)",(tuples[0],tuples[1],tuples[2],tuples[3],tuples[4]))
			conn.commit()
				
insert_vols(vols)			
	










		
			
			
			
